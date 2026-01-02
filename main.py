from datetime import datetime
import requests
import pandas as pd
from lxml import html, etree
import os
from openpyxl import load_workbook, Workbook
import time
import random

def append_df_to_excel(filename, df, sheet_name='Sheet1'):
    if not os.path.exists(filename):
        # 文件不存在：直接用 pandas 创建
        df.to_excel(filename, sheet_name=sheet_name, index=False)
    else:
        # 文件存在：用 openpyxl 打开并追加
        workbook = load_workbook(filename)
        if sheet_name not in workbook.sheetnames:
            worksheet = workbook.create_sheet(sheet_name)
            # 写入表头
            worksheet.append(df.columns.tolist())
        else:
            worksheet = workbook[sheet_name]

        # 追加数据行（逐行）
        for row in df.itertuples(index=False, name=None):
            worksheet.append(row)

        workbook.save(filename)

base_url = "http://www.unitednews.net.ph/"
cookies = {
    '_ga': 'GA1.1.17021160.1766027272',
    '_ga_R6Q15675G4': 'GS2.1.s1766060425$o3$g1$t1766060628$j60$l0$h0',
}
headers = {
    'Accept': '*/*',
    'Accept-Language': 'zh-CN,zh;q=0.9',
    'Cache-Control': 'no-cache',
    'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
    'Origin': 'http://www.unitednews.net.ph',
    'Pragma': 'no-cache',
    'Proxy-Connection': 'keep-alive',
    'Referer': 'http://www.unitednews.net.ph/category_page.php?category=6',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/143.0.0.0 Safari/537.36',
    'X-Requested-With': 'XMLHttpRequest',
    # 'Cookie': '_ga=GA1.1.17021160.1766027272; _ga_R6Q15675G4=GS2.1.s1766060425$o3$g1$t1766060628$j60$l0$h0',
}
data = {
    'limit': 10,
    'category': '6',
}

# 初始化结果列表
BATCH_SIZE = 100
batch_data = []  # 当前批次
output_file = "unitednews_articles.xlsx"

for data_size in range(705,801,5): # （205，801，5）  305
    data.update({"start":data_size})
    response = requests.post('http://www.unitednews.net.ph/getnews.php', cookies=cookies, headers=headers, data=data, verify=False)
    doc = html.fromstring(response.text)

    # 帖子链接
    url_list = doc.xpath("//a/@href")
    # 帖子标题
    title_list = doc.xpath("//div[@class=\"thumbs-title-container\"]/h5/text()")
    # 帖子发布时间
    time_list = doc.xpath("//small/text()")

    print(f"正在爬取第{(data_size + 5) / 5}批数据，共记{len(url_list)}条数据")
    print(url_list)
    # print(time_list)
    # break

    # 对齐数据：取最小长度，避免错位
    min_len = min(len(url_list), len(title_list), len(time_list))

    for i in range(min_len):
        url = url_list[i]
        title = title_list[i].strip()
        raw_time = time_list[i].strip()

        # 解析日期：格式为 "December 16, 2025"
        try:
            pub_time = datetime.strptime(raw_time, "%B %d, %Y")
        except ValueError as e:
            print(f" 日期解析失败: {raw_time} | 错误: {e}")
            pub_time = None  # 或用 pd.NaT

        # 构造详情页 URL
        try:
            post_id = url.split("=")[1]
            detail_url = base_url + url
            params = {"post": post_id}

            detail_resp = requests.get(
                detail_url,
                cookies=cookies,
                headers=headers,
                params=params,
                verify=False,timeout=10
            )
            detail_doc = html.fromstring(detail_resp.text)
            # 获取帖子内容
            content_nodes = detail_doc.xpath("/html/body/div[3]/div[2]/div[2]/div/div/div[2]/div/div[3]/p//text()")
            content = ''.join(content_nodes).strip()
        except Exception as e:
            print(f" 内容提取失败 for {url}: {e}")
            content = ""

        # 收集数据
        batch_data.append({
            "URL": detail_url,
            "Title": title,
            "Publish_Time": pub_time,  # ← 保留为 datetime 对象！
            "Content": content
        })
        time.sleep(random.uniform(0.5, 1))

        # 每满 BATCH_SIZE 条就保存
        if len(batch_data) >= BATCH_SIZE:
            df_batch = pd.DataFrame(batch_data)
            append_df_to_excel(output_file, df_batch)
            print(f" 已保存 {len(batch_data)} 条，累计约 {(data_size + 5) * 2} 条")
            batch_data = []  # 清空批次


# 循环结束后，保存剩余数据
if batch_data:
    df_batch = pd.DataFrame(batch_data)
    append_df_to_excel(output_file, df_batch)
    print(f" 最后一批 {len(batch_data)} 条已保存")

print(" 全部完成！")



